"""
求解问题
"""
import itertools
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import networkx as nx
from scipy.stats import rankdata
from tqdm import tqdm
from openpyxl import Workbook

from optimizer.problem_graph import graphProblem
from optimizer.aco import Colony, generate_routes_from_pheromone, generate_solution_from_routes
from optimizer.local_optimal import apply_local_optimal


def add_matrix_to_ws(ws, data):
    """
    将data中的数增添到ws现有数据区域下方，保持matrix的行列形状不变。
    :param ws: openpyxl的excel worksheet
    :param data: 一个numpy 二维数组或者矩阵
    """
    # 获取当前工作表的最大行数
    max_row = ws.max_row
    # 遍历data中的每一行
    for row_index, row in enumerate(data, start=1):
        # 遍历每一行中的每一个元素
        for col_index, value in enumerate(row, start=1):
            # 将元素写入到工作表中，位置是当前最大行数加上行索引
            ws.cell(row=max_row + row_index, column=col_index, value=value)
    

print("开始加载待求解问题")
DATA_PATH = r'D:\Development\code_commit_repo\vrp\dataset\test_data_100_nodes\data.pkl'
OUTPUT_PATH = r'D:\Development\code_commit_repo\vrp\dataset\test_data_100_nodes\result'
if not os.path.exists(OUTPUT_PATH):
    os.makedirs(OUTPUT_PATH)
    os.makedirs(os.path.join(OUTPUT_PATH, 'img'))
    os.makedirs(os.path.join(OUTPUT_PATH, 'img', 'solution'))
    os.makedirs(os.path.join(OUTPUT_PATH, 'img', 'pheromone'))

plm = graphProblem(data_path=DATA_PATH, output_path=OUTPUT_PATH)
plm.set_pheromones()
plm.show_graph()
nodes_size = len([f for f in plm.G.nodes if plm.G.nodes[f]['type'] == 'F'])
ALPHA = 1
BETA = 0.5
RHO = 0.05
GAMMA = 0.95
MAX_LOOP = 1000
NO_IMPRO_THRESHOLD = 50
MAX_PATH_LENGTH = 4
COLONY_SIZE = 100
PLOT_INTERVAL = 10
best_so_far_solution = None
best_so_far_cost = float('inf')
no_improvement_count = 0
pheromone = None
his_costs = []
his_solutions = []
his_pheromones = []
his_routes = []
gamma = GAMMA

for i in tqdm(range(MAX_LOOP), desc="当前成本: ", leave=False):

    colony = Colony(id=f'Colony|{i}',
                    k=COLONY_SIZE,
                    alpha=ALPHA,
                    beta=BETA,
                    gamma=GAMMA,
                    rho=RHO,
                    G=plm.G,
                    max_path_length=MAX_PATH_LENGTH,
                    fix_cost_func=plm.ftl_fix_cost,
                    pheromone=pheromone if pheromone is not None else plm.pheromone)
    colony.iterate()
    solution = colony.solution
    routes = colony.collect_routs()
    cost = plm.calculate_cost(solution_path=routes)
    his_costs.append(cost)
    his_pheromones.append(pheromone if pheromone is not None else plm.pheromone)
    his_solutions.append(solution)
    if cost < best_so_far_cost:
        opt_routes, opt_cost = apply_local_optimal(solution=routes,
                                                   cost_func=plm.calculate_cost,
                                                   nodes_size=nodes_size)
        if opt_cost < cost:
            cost = opt_cost
            solution = generate_solution_from_routes(solution, opt_routes)
            his_costs[-1] = cost
            his_solutions[-1] = solution
        his_routes.append(opt_routes)
        pheromone =  (1 - RHO)  * colony.pheromone + RHO * solution
        best_so_far_solution = solution
        best_so_far_cost = cost
        no_improvement_count = 0
    else:
        his_routes.append(routes)
        no_improvement_count += 1
        rank_new_cost = np.argsort(his_costs)[-1]
        pheromone =  (1 - RHO)  * pheromone + RHO * (rank_new_cost/len(his_costs) * solution
                                                     +  best_so_far_solution)

    if no_improvement_count >= NO_IMPRO_THRESHOLD:  # 如果大于NO_IMPRO_THRESHOLD的循环次数内没有改进，则停止
        break
    gamma = gamma * gamma
    if i % PLOT_INTERVAL == 0:
        result,solution = generate_routes_from_pheromone(plm.G, pheromone)
        plm.show_graph(highlights=solution, filename=f'img/solution/{i}.png')
        plm.show_graph(highlights=pheromone * 10, filename=f'img/pheromone/{i}.png')
    tqdm.write(f"当前成本: {cost}")

print("最终结果")
print("Best Sofar Solution:", best_so_far_solution)
print("Best Sofar Cost:", best_so_far_cost)
plm.show_graph(highlights=best_so_far_solution*2, filename="best_sofar_solution.png")
result,solution = generate_routes_from_pheromone(plm.G, pheromone)
plm.show_graph(highlights=solution*2, filename="final_solution.png")

print('最终方案的成本为：',plm.calculate_cost(result))

# 将结果输出到Excel文件中

wb = Workbook()
ws_solution = wb.active
ws_solution.title = 'Solution'

ws_solution.append([str(r) for r in result])
ws_pheromone = wb.create_sheet('Pheromone')
ws_solution.append(["Costs"])
ws_solution.append(his_costs)

ws_routes = wb.create_sheet('Pheromone')
ws_routes.title = 'Routes'
ws_routes.append(['Costs','Routes'])


for i, (p, s, rs, c) in enumerate(zip(his_pheromones, his_solutions, his_routes, his_costs)):
    ws_solution.append([f"Solution Iter {i}"])
    add_matrix_to_ws(ws_solution,s)
    ws_pheromone.append([f"Pheromone Iter {i}"])
    add_matrix_to_ws(ws_pheromone, p)
    ws_pheromone.append([f"Pheromone Iter {i}"])
    ws_routes.append([c]+[str(r) for r in rs])
wb.save(OUTPUT_PATH + r'\result.xlsx')


# Print best sofar cost
plt.plot(his_costs)
plt.scatter(his_costs.index(best_so_far_cost), best_so_far_cost, color='red')
plt.xlabel('Iteration')
plt.ylabel('Cost')
plt.title('Cost Over Iterations')
plt.show()